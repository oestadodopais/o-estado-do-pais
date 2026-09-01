#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
amostra-sonnet.py

Medicao cega de uma amostra do inventario de fontes do sitio O Estado do Pais.
Modelo: Claude Sonnet 5. Medicao feita em 01 de setembro de 2026.

O que faz: para cada uma das oito linhas de amostra-linhas.json (E3, T3, P1,
M2, S3, A1, D3, H4), contacta as fontes primarias (INE, Eurostat, ERSAR)
diretamente, sem consultar o trabalho de quem escreveu a linha, e compara o
que a fonte diz hoje com o que a linha afirma. Escreve o relatorio em
amostra-sonnet.md, na mesma pasta deste ficheiro.

Regras seguidas:
- User-Agent: OEstadoDoPais/medicao em todos os pedidos.
- Pedidos ao INE em serie, nunca em paralelo, com pelo menos 2 segundos de
  intervalo entre pedidos.
- Em HTTP 429, espera 30 segundos e tenta uma vez mais.
- O INE nao devolveu nenhum 429 durante esta medicao; devolveu, por vezes,
  tempo esgotado (60s) ou ligacao recusada. Para esses casos aplicou-se a
  mesma logica de segunda tentativa (ver funcao `fetch`).

Dependencias: biblioteca padrao do Python 3, mais `openpyxl` (para ler os
ficheiros .xlsx da ERSAR), o binario `curl` no PATH (usado para todos os
pedidos HTTP; ver nota abaixo) e o binario `pdftotext` (poppler-utils) no
PATH, para extrair texto do PDF do RASARP 2025.

Nota sobre `curl` em vez de `urllib`: numa primeira versao este script usava
`urllib.request` (biblioteca padrao) para os pedidos HTTP. Nesta maquina,
`urllib` falhou a verificacao TLS em praticamente todos os pedidos, com dois
erros diferentes consoante o anfitriao: "self-signed certificate in
certificate chain" (INE) e "unable to get local issuer certificate"
(Eurostat, ERSAR): o cafile por omissao do Python nesta instalacao
(/Library/Frameworks/Python.framework/Versions/3.14/etc/openssl/cert.pem)
nao valida estas cadeias. O `curl` desta mesma maquina (que usa o chaveiro
do sistema, via SecureTransport) validou sempre estes tres sitios sem
qualquer falha, ao longo de dezenas de pedidos feitos hoje. Por isso, e
porque foi de facto o que se usou na investigacao interativa que precedeu
este script, os pedidos HTTP aqui sao feitos via `curl` (subprocesso), nao
via `urllib`.
"""

import html
import json
import os
import re
import subprocess
import sys
import time
from datetime import datetime, timezone

UA = "OEstadoDoPais/medicao"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
AMOSTRA_JSON = os.path.join(SCRIPT_DIR, "amostra-linhas.json")
RELATORIO_MD = os.path.join(SCRIPT_DIR, "amostra-sonnet.md")
DOWNLOADS = "/tmp/amostra-sonnet-downloads"
os.makedirs(DOWNLOADS, exist_ok=True)

pedidos_log = []   # (url, hora_utc, status_ou_None)
falhas_log = []    # descricao textual de cada pedido que falhou na primeira tentativa


def agora():
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


_MARCA = "\n__HTTP_STATUS__:"


def _get(url, timeout):
    """Busca url com curl (segue redirecionamentos), devolvendo
    (status_http_ou_None, corpo_em_bytes, erro_ou_None)."""
    t0 = agora()
    try:
        proc = subprocess.run(
            ["curl", "-s", "-A", UA, "--max-time", str(int(timeout)), "-L",
             "-w", _MARCA + "%{http_code}", url],
            capture_output=True, timeout=timeout + 20,
        )
    except subprocess.TimeoutExpired:
        pedidos_log.append((url, t0, None))
        return None, b"", f"curl excedeu o tempo limite ({timeout}s)"
    out = proc.stdout
    i = out.rfind(_MARCA.encode())
    if i == -1:
        pedidos_log.append((url, t0, None))
        erro = proc.stderr.decode("utf-8", errors="replace").strip()[:300]
        return None, b"", f"curl sem marca de estado (returncode={proc.returncode}): {erro}"
    body = out[:i]
    status_str = out[i + len(_MARCA):].decode("ascii", errors="replace").strip()
    status = int(status_str) if status_str.isdigit() else None
    pedidos_log.append((url, t0, status))
    if status is None or status == 0:
        erro = proc.stderr.decode("utf-8", errors="replace").strip()[:300]
        return None, body, f"curl nao obteve resposta HTTP (returncode={proc.returncode}): {erro}"
    return status, body, None


def fetch(url, ine=False, timeout=60):
    """Busca um URL com o User-Agent da medicao.

    Para pedidos ao INE (ine=True), espera 2s no fim, sempre, para manter a
    serie com intervalo minimo de 2 segundos, nunca em paralelo.
    Em HTTP 429, espera 30s e tenta uma vez mais.
    Em falha de ligacao/tempo esgotado (sem resposta), espera 10s (ou 30s se
    for outra vez o INE, ja visto hoje a ser lento) e tenta uma vez mais.
    """
    status, body, err = _get(url, timeout)
    if status == 429:
        falhas_log.append(f"{url} -> HTTP 429; a aguardar 30s e a repetir uma vez")
        time.sleep(30)
        status, body, err = _get(url, timeout)
    elif status is None:
        falhas_log.append(f"{url} -> sem resposta ({err}); a repetir uma vez")
        time.sleep(10)
        status, body, err = _get(url, max(timeout, 90))
    if ine:
        time.sleep(2)
    return status, body, err


def texto_de_html(raw_bytes):
    t = raw_bytes.decode("utf-8", errors="replace")
    t = re.sub("<[^>]+>", " ", t)
    t = html.unescape(t)
    t = t.replace("\xa0", " ").replace("​", "")  # espaco fixo e espaco de largura zero -> normalizados
    t = re.sub(r"[ \t]+", " ", t)
    t = re.sub(r"\n\s*\n+", "\n", t)
    return t


def contem_tudo(texto, *fragmentos):
    """Verdadeiro se TODOS os fragmentos aparecem no texto (cada um deve ser
    contiguo no texto de origem; usa-se para confirmar citacoes verbatim)."""
    return all(f in texto for f in fragmentos)


class Linha:
    """Acumula os resultados de uma linha (celula, afirma, fonte, veredito)."""

    def __init__(self, id_):
        self.id = id_
        self.resultados = []

    def bate(self, celula, afirma, fonte):
        self.resultados.append((celula, afirma, fonte, "bate"))

    def nao_bate(self, celula, afirma, fonte):
        self.resultados.append((celula, afirma, fonte, "NAO BATE"))

    def na(self, celula, afirma, motivo):
        self.resultados.append((celula, afirma, f"(nao consultado: {motivo})", "nao testavel daqui"))


# ---------------------------------------------------------------------------
# E3 - Eurostat, gov_10dd_edpt1 (divida publica)
# ---------------------------------------------------------------------------

def checar_E3(linha_json):
    L = Linha("E3")
    url = linha_json["url_maquina"]
    status, body, err = fetch(url)
    if status != 200:
        L.na("(todas)", "-", f"url_maquina devolveu {status}/{err}")
        return L
    d = json.loads(body)

    L.na("publicador_primario", "INE (notificacao do PDE), como em E2",
         "o ESMS (gov_10dd_esms.htm) descreve o mecanismo geral (os Estados-Membros reportam via as suas "
         "institutos nacionais, Reg. 479/2009) mas nao nomeia Portugal/INE; a linha remete para E2, fora "
         "da amostra")

    # definicao (verbatim, confirmada contra o ESMS do dataset)
    st_esms, body_esms, _ = fetch("https://ec.europa.eu/eurostat/cache/metadata/en/gov_10dd_esms.htm")
    t_esms = texto_de_html(body_esms) if st_esms == 200 else ""
    definicao_esperada = ("The government debt is defined as the total consolidated gross debt at "
                           "nominal (face) value at the end of the year in the following categories of "
                           "government liabilities (as defined in ESA 2010): currency and deposits (AF.2), "
                           "debt securities (AF.3) and loans (AF.4).")
    ok = definicao_esperada in t_esms
    (L.bate if ok else L.nao_bate)("definicao", definicao_esperada[:90] + "...",
                                    "confirmado verbatim no ESMS" if ok else "texto nao encontrado no ESMS")

    # serie_desde / ultimo_periodo / publicado_em / excerto / valor_recente
    idx = d["dimension"]["time"]["category"]["index"]
    anos = sorted(idx, key=lambda k: idx[k])
    inv = {v: k for k, v in idx.items()}
    serie = {inv[int(p)]: v for p, v in d["value"].items()}
    primeiro, ultimo = anos[0], anos[-1]
    L.bate("serie_desde", "1995", f"primeiro ano na serie = {primeiro}") if primeiro == "1995" else \
        L.nao_bate("serie_desde", "1995", f"primeiro ano na serie = {primeiro}")
    L.bate("ultimo_periodo", "2025", f"ultimo ano na serie = {ultimo}") if ultimo == "2025" else \
        L.nao_bate("ultimo_periodo", "2025", f"ultimo ano na serie = {ultimo}")

    updated = d.get("updated", "")
    ok = updated.startswith("2026-04-22")
    (L.bate if ok else L.nao_bate)("publicado_em", "2026-04-22", f"updated = {updated}")

    ok = serie.get("2023") == 96.9 and serie.get("2024") == 93.5 and serie.get("2025") == 89.7
    (L.bate if ok else L.nao_bate)("excerto",
                                    "2023: 96.9 ; 2024: 93.5 ; 2025: 89.7",
                                    f"2023: {serie.get('2023')} ; 2024: {serie.get('2024')} ; 2025: {serie.get('2025')}")
    ok = serie.get("2025") == 89.7
    (L.bate if ok else L.nao_bate)("valor_recente", "89,7% do PIB em 2025", f"2025 = {serie.get('2025')} (unit=PC_GDP)")

    # periodicidade: "Dados trimestrais, difusao semestral (abril e outubro)"
    freq = list(d["dimension"]["freq"]["category"]["label"].values())[0]
    freq_ok = freq == "Annual"  # a linha diz "trimestrais"; a fonte diz freq=A=Annual
    disseminacao_ok = "Bi-annual" in t_esms and "before 1 April and before 1 October" in t_esms
    if (not freq_ok) and disseminacao_ok:
        L.nao_bate("periodicidade",
                    "Dados trimestrais, difusao semestral (abril e outubro)",
                    f"dimensao freq = '{freq}' (Annual, nao trimestral); ESMS 9. Frequency of dissemination = "
                    "'Bi-annual', reporte 'before 1 April and before 1 October' (confirma so a parte da difusao semestral)")
    else:
        L.bate("periodicidade", "Dados trimestrais, difusao semestral (abril e outubro)", f"freq={freq}")

    # concelho: "nao" -> dataset sem nivel geografico abaixo de pais
    st2, body2, _ = fetch("https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/"
                           "gov_10dd_edpt1?format=JSON&lang=EN&unit=PC_GDP&sector=S13&na_item=GD&time=2025")
    if st2 == 200:
        geo_all = json.loads(body2)["dimension"]["geo"]["category"]["index"]
        # pergunta relevante: ha alguma repartição infranacional para Portugal? um codigo
        # NUTS portugues seria "PT" + sufixo (ex.: PT11, PT150); nenhum codigo agregado
        # (EU27_2020, EA19/20/21, etc.) comeca por "PT".
        pt_sub = [g for g in geo_all if g.startswith("PT") and g != "PT"]
        ok = len(pt_sub) == 0
        (L.bate if ok else L.nao_bate)("concelho", "nao",
                                        f"{len(geo_all)} posicoes geograficas no total; nenhuma comeca por 'PT' alem "
                                        f"do proprio 'PT' (sub-codigos PT encontrados: {pt_sub})")
    else:
        L.na("concelho", "nao", f"pedido sem filtro geo falhou ({st2})")

    # licenca
    st3, body3, _ = fetch("https://ec.europa.eu/eurostat/web/main/help/copyright-notice")
    t3 = texto_de_html(body3) if st3 == 200 else ""
    frag1 = ("The copyright for the editorial content of this website, which is owned by the EU, is "
             "licensed under the Creative Commons Attribution 4.0 International licence")
    frag2 = ("Reuse of statistical data, metadata, publications, and other dissemination tools published "
             "on this website for commercial or non-commercial purposes is authorised provided the source "
             "is acknowledged")
    ok = contem_tudo(t3, frag1, frag2)
    (L.bate if ok else L.nao_bate)("licenca", "(texto EU/Eurostat sobre direitos de autor)",
                                    "confirmado em ec.europa.eu/eurostat/web/main/help/copyright-notice" if ok
                                    else "texto nao encontrado na pagina de copyright")

    return L


# ---------------------------------------------------------------------------
# Auxiliares para indicadores do INE (pindica.jsp / pindicaMeta.jsp / minfo.jsp)
# ---------------------------------------------------------------------------

def ine_meta(varcd):
    url = f"https://www.ine.pt/ine/json_indicador/pindicaMeta.jsp?varcd={varcd}&lang=PT"
    status, body, err = fetch(url, ine=True)
    if status != 200:
        return None, url, status
    return json.loads(body)[0], url, status


def ine_minfo(varcd):
    url = f"https://www.ine.pt/bddXplorer/htdocs/minfo.jsp?var_cd={varcd}&lingua=PT"
    status, body, err = fetch(url, ine=True)
    if status != 200:
        return "", url, status
    return texto_de_html(body), url, status


def ine_data(varcd, dim1, dim2=None, dim3=None, dim4=None):
    url = f"https://www.ine.pt/ine/json_indicador/pindica.jsp?op=2&varcd={varcd}&Dim1={dim1}"
    if dim2:
        url += f"&Dim2={dim2}"
    if dim3:
        url += f"&Dim3={dim3}"
    if dim4:
        url += f"&Dim4={dim4}"
    url += "&lang=PT"
    status, body, err = fetch(url, ine=True)
    if status != 200:
        return None, url, status
    return json.loads(body)[0], url, status


_LICENCA_INE_CACHE = None


def licenca_ine_texto():
    """Busca uma so vez a pagina de termos do INE que traz o texto da licenca
    CC BY 4.0 (varias linhas citam o mesmo texto; evita repetir o pedido)."""
    global _LICENCA_INE_CACHE
    if _LICENCA_INE_CACHE is None:
        st, body, _ = fetch("https://www.ine.pt/xportal/xmain?xpid=INE&xpgid=ine_pufs_termos&contexto=up&xlang=pt",
                             ine=True)
        _LICENCA_INE_CACHE = texto_de_html(body) if st == 200 else ""
    return _LICENCA_INE_CACHE


def niveis_geo(meta):
    dims = meta["Dimensoes"]["Categoria_Dim"][0]
    niveis = {}
    for k, v in dims.items():
        if k.startswith("Dim_Num2"):
            n = v[0]["categ_nivel"]
            niveis[n] = niveis.get(n, 0) + 1
    return niveis


def valores_por_geo(data, ano):
    linhas = data["Dados"].get(ano, [])
    return {r["geocod"]: r["ind_string"] for r in linhas}


def valores_por_geo_e_tipo(data, ano):
    linhas = data["Dados"].get(ano, [])
    out = {}
    for r in linhas:
        out[(r["geocod"], r.get("dim_3_t"))] = r["ind_string"]
    return out


# ---------------------------------------------------------------------------
# T3 - Ganho medio mensal (INE 0012656)
# ---------------------------------------------------------------------------

def checar_T3(linha_json):
    L = Linha("T3")
    meta, url_meta, st = ine_meta("0012656")
    if meta is None:
        L.na("(todas)", "-", f"pindicaMeta falhou ({st})")
        return L
    minfo, url_minfo, st_m = ine_minfo("0012656")
    data, url_data, st_d = ine_data("0012656", "S7A2024", dim2="PT,1C40705")

    fonte = meta.get("Fonte") or (re.search(r"Fonte\s*\n?\s*([^\n]+)", minfo).group(1).strip() if "Fonte" in minfo else None)
    ok = fonte is not None and "MTSSS/GEP, Quadros de pessoal" in fonte
    (L.bate if ok else L.nao_bate)("publicador_primario", "Fonte: MTSSS/GEP, Quadros de pessoal", f"Fonte = {fonte!r}")

    ganho_ok = "Montante ilíquido em dinheiro e/ou géneros pago ao trabalhador com caráter regular" in minfo
    nota_ok = meta.get("Nota", "") == ("Os dados referem-se a trabalhadores por conta de outrem a tempo completo "
                                        "com remuneração completa.")
    (L.bate if (ganho_ok and nota_ok) else L.nao_bate)(
        "definicao", "GANHO: Montante iliquido... || Nota: trabalhadores por conta de outrem a tempo completo...",
        f"GANHO confirmado={ganho_ok}; Nota confirmada={nota_ok} (Nota atual={meta.get('Nota')!r})")

    ok = meta.get("PrimeiroPeriodo") == "2021"
    (L.bate if ok else L.nao_bate)("serie_desde", "2021", f"PrimeiroPeriodo={meta.get('PrimeiroPeriodo')}")

    ok = meta.get("Periodic") == "Anual"
    (L.bate if ok else L.nao_bate)("periodicidade", "Anual", f"Periodic={meta.get('Periodic')}")

    ok = meta.get("UltimoPeriodo") == "2024"
    (L.bate if ok else L.nao_bate)("ultimo_periodo", "2024", f"UltimoPeriodo={meta.get('UltimoPeriodo')}")

    ok = meta.get("DataUltimaAtualizacao") == "2026-03-27"
    (L.bate if ok else L.nao_bate)("publicado_em", "2026-03-27", f"DataUltimaAtualizacao={meta.get('DataUltimaAtualizacao')}")

    niveis = niveis_geo(meta)
    ok = niveis == {"1": 1, "2": 3, "3": 9, "4": 26, "5": 308}
    (L.bate if ok else L.nao_bate)("concelho", "sim, os 308 (1+3+9+26+308 categorias)", f"niveis={niveis}")

    t_lic = licenca_ine_texto()
    frag = ("A informação estatística disponível neste Portal é gratuita e pode ser usada e reutilizada "
            "pelos utilizadores de acordo com a licença Creative Commons CC BY Atribuição 4.0")
    ok = frag in t_lic
    (L.bate if ok else L.nao_bate)("licenca", frag, "confirmado em ine_pufs_termos" if ok else "nao encontrado")

    if data:
        vals = valores_por_geo(data, "2024")
        ok = vals.get("PT") == "1 576,0" and vals.get("1C40705") == "1 484,5"
        (L.bate if ok else L.nao_bate)("excerto / valor_recente",
                                        "Portugal 2024: 1 576,0 ; Evora 2024: 1 484,5",
                                        f"Portugal={vals.get('PT')} ; Evora={vals.get('1C40705')}")
    else:
        L.na("excerto / valor_recente", "-", f"pindica falhou ({st_d})")

    return L


# ---------------------------------------------------------------------------
# P1 - Populacao residente (INE 0012918, irmao 0012917)
# ---------------------------------------------------------------------------

def checar_P1(linha_json):
    L = Linha("P1")
    meta, _, st = ine_meta("0012918")
    if meta is None:
        L.na("(todas)", "-", f"pindicaMeta falhou ({st})")
        return L
    minfo, _, _ = ine_minfo("0012918")
    data, _, st_d = ine_data("0012918", "S7A2025", dim2="PT,1C40705", dim3="T", dim4="T")
    data_irmao, _, st_i = ine_data("0012917", "S7A2025", dim2="PT,1C40705", dim3="T", dim4="T")

    fonte_ok = "INE, Estimativas anuais da população residente" in minfo
    (L.bate if fonte_ok else L.nao_bate)("publicador_primario", "INE, Estimativas anuais da população residente",
                                          f"encontrado no minfo={fonte_ok}")

    def_ok = ("Conjunto de pessoas que, independentemente de estarem presentes ou ausentes num determinado "
              "alojamento no momento de observação" in minfo)
    formula_ok = "Valor estimado" in minfo
    (L.bate if (def_ok and formula_ok) else L.nao_bate)(
        "definicao", "POPULACAO RESIDENTE: ... Formula: Valor estimado",
        f"definicao confirmada={def_ok}; formula confirmada={formula_ok}")

    ok = meta.get("PrimeiroPeriodo") == "2021"
    (L.bate if ok else L.nao_bate)("serie_desde", "2021", f"PrimeiroPeriodo={meta.get('PrimeiroPeriodo')}")

    ok = meta.get("Periodic") == "Anual"
    (L.bate if ok else L.nao_bate)("periodicidade", "Anual", f"Periodic={meta.get('Periodic')}")

    ok = meta.get("UltimoPeriodo") == "2025"
    (L.bate if ok else L.nao_bate)("ultimo_periodo", "2025", f"UltimoPeriodo={meta.get('UltimoPeriodo')}")

    # publicado_em: a linha afirma 2026-05-22; a fonte (hoje) diz 2026-06-22
    dua = meta.get("DataUltimaAtualizacao")
    ok = dua == "2026-05-22"
    L.nao_bate("publicado_em", "2026-05-22 (DataUltimaAtualizacao)", f"DataUltimaAtualizacao={dua}") if not ok else \
        L.bate("publicado_em", "2026-05-22", f"DataUltimaAtualizacao={dua}")

    # a propria linha cita o destaque a abrir com "22 de junho de 2026" - conferir na fonte do destaque
    st_dest, body_dest, _ = fetch(
        "https://www.ine.pt/xportal/xmain?xpid=INE&xpgid=ine_destaques&DESTAQUESdest_boui=770295679&DESTAQUESmodo=2&xlang=pt",
        ine=True)
    t_dest = texto_de_html(body_dest) if st_dest == 200 else ""
    frag_dest = ("Em 31 de dezembro de 2025, a população residente em Portugal foi estimada em 11 424 031 "
                 "pessoas, o que corresponde a um aumento de 36 809 pessoas relativamente a 2024 (0,32%).")
    ok = "22 de junho de 2026" in t_dest and frag_dest in t_dest
    (L.bate if ok else L.nao_bate)("publicado_em (destaque)",
                                    "destaque abre com 22 de junho de 2026 / texto do resumo",
                                    f"confirmado={ok}")

    niveis = niveis_geo(meta)
    ok = niveis == {"1": 1, "2": 3, "3": 9, "4": 26, "5": 308}
    (L.bate if ok else L.nao_bate)("concelho", "sim, os 308 (347 categorias: 1+3+9+26+308)", f"niveis={niveis}")

    t_lic = licenca_ine_texto()
    frag = "licença Creative Commons CC BY Atribuição 4.0"
    ok = frag in t_lic
    (L.bate if ok else L.nao_bate)("licenca", frag, "confirmado" if ok else "nao encontrado")

    if data and data_irmao:
        vals = valores_por_geo(data, "2025")
        vals_irmao = valores_por_geo(data_irmao, "2025")
        ok = (vals.get("PT") == "11 424 031" and vals.get("1C40705") == "58 567" and
              vals_irmao.get("PT") == "11 424 031" and vals_irmao.get("1C40705") == "58 567")
        (L.bate if ok else L.nao_bate)(
            "excerto / valor_recente",
            "Portugal 2025: 11 424 031 ; Evora 2025: 58 567 ; identicos em 0012918 e 0012917",
            f"0012918: PT={vals.get('PT')} Evora={vals.get('1C40705')} ; "
            f"0012917: PT={vals_irmao.get('PT')} Evora={vals_irmao.get('1C40705')}")
    else:
        L.na("excerto / valor_recente", "-", f"pindica falhou ({st_d}/{st_i})")

    return L


# ---------------------------------------------------------------------------
# M2 - Populacao estrangeira com estatuto legal de residente (INE 0013220)
# ---------------------------------------------------------------------------

def checar_M2(linha_json):
    L = Linha("M2")
    meta, _, st = ine_meta("0013220")
    if meta is None:
        L.na("(todas)", "-", f"pindicaMeta falhou ({st})")
        return L
    minfo, _, _ = ine_minfo("0013220")
    data, _, st_d = ine_data("0013220", "S7A2023", dim2="PT,1C40705", dim3="T", dim4="T")

    fonte_ok = "INE, População estrangeira com estatuto legal de residente" in minfo
    (L.bate if fonte_ok else L.nao_bate)("publicador_primario",
                                          "Fonte: INE, População estrangeira com estatuto legal de residente",
                                          f"encontrado no minfo={fonte_ok}")

    L.na("publicador_primario (doc. metodologico 443 / I000748)", "-",
         "catalogo de metodologia do INE nao consultado; nao referenciado nos enderecos da linha")

    def_ok = ("Conjunto de pessoas de nacionalidade não portuguesa com autorização ou cartão de residência"
              in minfo)
    (L.bate if def_ok else L.nao_bate)("definicao", "POPULACAO ESTRANGEIRA COM ESTATUTO LEGAL DE RESIDENTE: ...",
                                        f"confirmado={def_ok}")

    ok = meta.get("PrimeiroPeriodo") == "2021"
    (L.bate if ok else L.nao_bate)("serie_desde", "2021", f"PrimeiroPeriodo={meta.get('PrimeiroPeriodo')}")

    ok = meta.get("Periodic") == "Anual"
    (L.bate if ok else L.nao_bate)("periodicidade", "Anual", f"Periodic={meta.get('Periodic')}")

    ok = meta.get("UltimoPeriodo") == "2023"
    (L.bate if ok else L.nao_bate)("ultimo_periodo", "2023", f"UltimoPeriodo={meta.get('UltimoPeriodo')}")

    ok = meta.get("DataUltimaAtualizacao") == "2024-09-20"
    (L.bate if ok else L.nao_bate)("publicado_em", "2024-09-20", f"DataUltimaAtualizacao={meta.get('DataUltimaAtualizacao')}")

    niveis = niveis_geo(meta)
    ok = niveis.get("5") == 308
    (L.bate if ok else L.nao_bate)("concelho", "sim, os 308, para 2021 a 2023", f"niveis={niveis}")

    t_lic = licenca_ine_texto()
    ok = "licença Creative Commons CC BY Atribuição 4.0" in t_lic
    (L.bate if ok else L.nao_bate)("licenca", "CC BY Atribuicao 4.0", "confirmado" if ok else "nao encontrado")

    if data:
        vals = valores_por_geo(data, "2023")
        ok = vals.get("PT") == "1 044 238" and vals.get("1C40705") == "3 684"
        (L.bate if ok else L.nao_bate)("excerto / valor_recente",
                                        "Portugal 2023: 1 044 238 ; Evora 2023: 3 684",
                                        f"Portugal={vals.get('PT')} ; Evora={vals.get('1C40705')}")
    else:
        L.na("excerto / valor_recente", "-", f"pindica falhou ({st_d})")

    return L


# ---------------------------------------------------------------------------
# S3 - Pensionistas e valor medio das pensoes (INE 0014534 e 0014532)
# ---------------------------------------------------------------------------

def checar_S3(linha_json):
    L = Linha("S3")
    meta_a, _, st_a = ine_meta("0014534")
    meta_b, _, st_b = ine_meta("0014532")
    if meta_a is None or meta_b is None:
        L.na("(todas)", "-", f"pindicaMeta falhou ({st_a}/{st_b})")
        return L
    minfo_a, _, _ = ine_minfo("0014534")
    minfo_b, _, _ = ine_minfo("0014532")
    data_a, _, st_da = ine_data("0014534", "S7A2025", dim2="PT")
    data_b, _, st_db = ine_data("0014532", "S7A2025", dim2="PT")

    fonte_ok = "Instituto de Informática" in minfo_a
    (L.bate if fonte_ok else L.nao_bate)("publicador_primario", "Fonte: Instituto de Informática",
                                          f"confirmado={fonte_ok}")

    pens_ok = "PENSIONISTA : Titular de uma prestação pecuniária por invalidez, velhice, doença profissional ou morte." in minfo_a
    formula_ok = "Valor das pensões da segurança social/ Pensionistas da segurança social" in minfo_b
    obs = ("A partir de janeiro de 2017, contabiliza-se o número de pensionistas com pagamento de pensão no "
           "âmbito da Segurança Social: Regime Geral, Regime Não Contributivo e Equiparado, Regime Especial "
           "de Segurança Social das Atividades Agrícolas. Os dados dizem respeito às pensões pagas pela "
           "Segurança Social.")
    obs_ok = obs in minfo_a and obs in minfo_b
    (L.bate if (pens_ok and formula_ok and obs_ok) else L.nao_bate)(
        "definicao", "PENSIONISTA: ...; Formula valor medio: ...; Observacoes: ...",
        f"PENSIONISTA={pens_ok}; formula={formula_ok}; observacoes={obs_ok}")

    ok = meta_a.get("PrimeiroPeriodo") == "2017" and meta_b.get("PrimeiroPeriodo") == "2017"
    (L.bate if ok else L.nao_bate)("serie_desde", "2017 em ambos",
                                    f"0014534={meta_a.get('PrimeiroPeriodo')} ; 0014532={meta_b.get('PrimeiroPeriodo')}")

    ok = meta_a.get("Periodic") == "Anual" and meta_b.get("Periodic") == "Anual"
    (L.bate if ok else L.nao_bate)("periodicidade", "Anual", f"{meta_a.get('Periodic')}/{meta_b.get('Periodic')}")

    ok = meta_a.get("UltimoPeriodo") == "2025" and meta_b.get("UltimoPeriodo") == "2025"
    (L.bate if ok else L.nao_bate)("ultimo_periodo", "2025",
                                    f"{meta_a.get('UltimoPeriodo')}/{meta_b.get('UltimoPeriodo')}")

    ok = meta_a.get("DataUltimaAtualizacao") == "2026-08-28" and meta_b.get("DataUltimaAtualizacao") == "2026-08-28"
    (L.bate if ok else L.nao_bate)("publicado_em", "28/08/2026",
                                    f"{meta_a.get('DataUltimaAtualizacao')}/{meta_b.get('DataUltimaAtualizacao')}")

    niveis = niveis_geo(meta_a)
    ok = niveis == {"1": 1, "2": 3, "3": 9, "4": 26, "5": 308}
    (L.bate if ok else L.nao_bate)("concelho", "sim, os 308 (347 categorias)", f"niveis={niveis}")

    st_termos, body_termos, _ = fetch("https://www.ine.pt/xportal/xmain?xpid=INE&xpgid=ine_princ_termos", ine=True)
    t_termos = texto_de_html(body_termos) if st_termos == 200 else ""
    ok = st_termos == 200 and "O Servidor encontra-se em serviço de manutenção" in t_termos
    (L.bate if ok else L.nao_bate)("licenca ([verify]: pagina de termos)",
                                    "xpgid=ine_princ_termos devolve 200 com corpo de manutencao",
                                    f"status={st_termos}; corpo confirma manutencao={ok}")

    if data_a and data_b:
        pens = {(r["dim_3_t"]): r["ind_string"] for r in data_a["Dados"].get("2025", [])}
        medio = {(r["dim_3_t"]): r["ind_string"] for r in data_b["Dados"].get("2025", [])}
        ok = (pens.get("Total") == "2 922 353" and pens.get("Velhice") == "2 030 992" and
              pens.get("Sobrevivência") == "724 023" and pens.get("Invalidez") == "167 338" and
              medio.get("Total") == "8 066" and medio.get("Velhice") == "9 519" and
              medio.get("Invalidez") == "7 043" and medio.get("Sobrevivência") == "4 228")
        (L.bate if ok else L.nao_bate)(
            "excerto / valor_recente",
            "Pensionistas PT 2025: Total 2 922 353, Velhice 2 030 992, Sobrevivencia 724 023, Invalidez 167 338; "
            "Valor medio: Total 8 066, Velhice 9 519, Invalidez 7 043, Sobrevivencia 4 228",
            f"pensionistas={pens} ; valor_medio={medio}")
    else:
        L.na("excerto / valor_recente", "-", f"pindica falhou ({st_da}/{st_db})")

    return L


# ---------------------------------------------------------------------------
# A1 - Agua nao faturada (ERSAR, RASARP 2025)
# ---------------------------------------------------------------------------

def checar_A1(linha_json):
    L = Linha("A1")
    url_idx = linha_json["url_maquina"].split(" ; ")[0].strip()
    status, body, err = fetch(url_idx)
    if status != 200:
        L.na("(todas)", "-", f"indice RASARP falhou ({status}/{err})")
        return L
    idx = json.loads(body)
    items = idx.get("items", [])
    titulos = [it["titulo"] for it in items]
    ok = titulos[0] == "RASARP 2025" and titulos[-1] == "RASARP 2004"
    (L.bate if ok else L.nao_bate)("serie_desde", "edicoes anuais do RASARP desde 2004 (RASARP 2004 a RASARP 2025)",
                                    f"primeiro item={titulos[0]!r} ; ultimo item={titulos[-1]!r} ; total={len(items)}")

    rasarp2025_v1 = next((it for it in items if it["titulo"] == "RASARP 2025" and "Volume 1" in it["subtitulo"]), None)
    if rasarp2025_v1 is None:
        L.na("(publicado_em / excertos do PDF)", "-", "nao encontrei RASARP 2025 Volume 1 no indice")
        return L
    pdf_url = next(f["path"] for f in rasarp2025_v1["files"] if f["originalName"].lower().endswith(".pdf"))
    xlsx_urls = {f["originalName"]: f["path"] for f in rasarp2025_v1["files"] if f["originalName"].endswith(".xlsx")}

    # PDF do RASARP 2025 Volume 1
    st_pdf, body_pdf, err_pdf = fetch(pdf_url, timeout=120)
    if st_pdf != 200:
        L.na("(publicado_em / excertos do PDF)", "-", f"download do PDF falhou ({st_pdf}/{err_pdf})")
        txt = ""
    else:
        pdf_path = os.path.join(DOWNLOADS, "rasarp2025_vol1.pdf")
        txt_path = os.path.join(DOWNLOADS, "rasarp2025_vol1.txt")
        with open(pdf_path, "wb") as fh:
            fh.write(body_pdf)
        subprocess.run(["pdftotext", "-layout", pdf_path, txt_path], check=True)
        with open(txt_path, encoding="utf-8", errors="replace") as fh:
            txt = fh.read()

    ok = "26 de fevereiro de 2026" in txt and "978-989-8360-49-6" in txt
    (L.bate if ok else L.nao_bate)("publicado_em", "26 de fevereiro de 2026; ISBN 978-989-8360-49-6",
                                    f"confirmado={ok}")

    def_esperada_frags = [
        "Pretende-se avaliar o nível de sustentabilidade da gestão do",
        "perdas económicas correspondentes à água que, apesar de",
        "não chega a ser faturada aos utilizadores.",
        "O indicador é definido como a percentagem de água entrada",
        "gestoras de sistemas em alta e em baixa).",
    ]
    ok = contem_tudo(txt, *def_esperada_frags)
    (L.bate if ok else L.nao_bate)("definicao", "Pretende-se avaliar o nivel de sustentabilidade... (verbatim)",
                                    f"todos os fragmentos encontrados={ok}")

    ok = "Mudança de geração" in txt and "Quadro 51. AA08 baixa" in txt
    (L.bate if ok else L.nao_bate)("serie_desde (quadro de evolucao 2020-2024)",
                                    "quadro de evolucao 2020-2024 com aviso 'Mudanca de geracao'", f"confirmado={ok}")

    ok = ("no Quadro 304 é de 284 em vez de 278" in txt)
    (L.bate if ok else L.nao_bate)("concelho (nota sobre Quadro 304)",
                                    "nota: total de concelhos no Quadro 304 e 284 em vez de 278", f"confirmado={ok}")

    frag_q48 = ["Água entrada no sistema", "652 687 845 m3/ano", "34 057 068 m3/ano", "5,2 %"]
    frag_q49 = ["846 741 648 m3/ano", "224 028 933 m3/ano", "26,5 %"]
    ok = contem_tudo(txt, *frag_q48) and contem_tudo(txt, *frag_q49)
    (L.bate if ok else L.nao_bate)(
        "excerto / valor_recente (Quadros 48 e 49)",
        "alta: 5,2% (34 057 068 m3/652 687 845 m3) ; baixa: 26,5% (224 028 933 m3/846 741 648 m3)",
        f"confirmado={ok}")

    # licenca (avisos legais do ersar.pt)
    st_lic, body_lic, _ = fetch("https://www.ersar.pt/pt_avisos-legais.html")
    t_lic = texto_de_html(body_lic) if st_lic == 200 else ""
    frag_lic1 = ("O conteúdo integral deste sítio, incluindo textos, gráficos, imagens, sons e quaisquer "
                 "outras informações, são propriedade da ERSAR")
    frag_lic2 = "Nota de direitos reservados a ​“© ERSAR AAAA. Todos os direitos Reservados."
    ok = frag_lic1 in t_lic and ("ERSAR AAAA" in t_lic)
    (L.bate if ok else L.nao_bate)("licenca", "texto de direitos de autor da ERSAR (avisos legais)", f"confirmado={ok}")

    # AnexoX.xlsx: 43661 linhas, 352 entidades, Aguas da Batalha AA08b = 26
    try:
        import openpyxl
        st_x, body_x, err_x = fetch(xlsx_urls["AnexoX-Relatorio_Indicadores_Dados_2024.xlsx"], timeout=120)
        if st_x == 200:
            xlsx_path = os.path.join(DOWNLOADS, "AnexoX.xlsx")
            with open(xlsx_path, "wb") as fh:
                fh.write(body_x)
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb["Indicadores_2024"]
            n_linhas = ws.max_row - 1
            empresas = set()
            achado = None
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                empresas.add(row[0])
                if row[0] == "Águas da Batalha" and row[4] == "AA08b":
                    achado = row
            ok = (n_linhas == 43661 and len(empresas) == 352 and achado is not None and achado[6] == 26)
            (L.bate if ok else L.nao_bate)(
                "url_maquina (AnexoX, linhas) / concelho (352 entidades) / excerto (Aguas da Batalha)",
                "43 661 linhas ; 352 entidades no ficheiro ; Aguas da Batalha AA08b = 26%",
                f"linhas={n_linhas} ; entidades={len(empresas)} ; linha encontrada={achado}")
        else:
            L.na("AnexoX.xlsx", "-", f"download falhou ({st_x})")
    except ImportError:
        L.na("AnexoX.xlsx", "-", "openpyxl nao disponivel neste ambiente")

    # AnexoVII.xlsx: 278 concelhos distintos na matriz de avaliacao AA baixa
    try:
        import openpyxl
        st_v, body_v, err_v = fetch(xlsx_urls["AnexoVII-Matriz_AQS_AA.xlsx"], timeout=60)
        if st_v == 200:
            v_path = os.path.join(DOWNLOADS, "AnexoVII.xlsx")
            with open(v_path, "wb") as fh:
                fh.write(body_v)
            wb2 = openpyxl.load_workbook(v_path, read_only=True, data_only=True)
            ws2 = wb2["Panorama Nacional AA_BAIXA"]
            concelhos = {row[2] for row in ws2.iter_rows(min_row=9, values_only=True) if row[2]}
            ok = len(concelhos) == 278
            (L.bate if ok else L.nao_bate)("concelho (matriz Anexo VII)",
                                            "278 concelhos distintos (Portugal continental)",
                                            f"concelhos distintos encontrados={len(concelhos)}")
        else:
            L.na("AnexoVII.xlsx", "-", f"download falhou ({st_v})")
    except ImportError:
        pass

    return L


# ---------------------------------------------------------------------------
# D3 - Taxa bruta de pre-escolarizacao (INE 0012616 + Eurostat educ_uoe_enra21)
# ---------------------------------------------------------------------------

def checar_D3(linha_json):
    L = Linha("D3")
    meta, _, st = ine_meta("0012616")
    if meta is None:
        L.na("(parte INE)", "-", f"pindicaMeta falhou ({st})")
    else:
        minfo, _, _ = ine_minfo("0012616")
        data, _, st_d = ine_data("0012616", "7A2025_3", dim2="PT,1,1C40705", dim3="T")

        fonte_ok = "Direção-Geral de Estatísticas da Educação e Ciência" in minfo
        (L.bate if fonte_ok else L.nao_bate)("publicador_primario (INE->DGEEC)",
                                              "Fonte: Direção-Geral de Estatísticas da Educação e Ciência",
                                              f"confirmado={fonte_ok}")

        formula_ok = "(Crianças inscritas na educação pré-escolar/ População residente com idade entre 3 a 5 anos)*100" in minfo
        (L.bate if formula_ok else L.nao_bate)("definicao (formula INE)", "formula INE verbatim", f"confirmado={formula_ok}")
        L.na("definicao (DGEEC, Designacao/Definicao Taxa Bruta de Escolarizacao)", "-",
             "sem endereco DGEEC nesta linha; nao consta na metainformacao do INE consultada")

        ok = meta.get("PrimeiroPeriodo") == "2014 / 2015"
        (L.bate if ok else L.nao_bate)("serie_desde (INE)", "2014 / 2015", f"PrimeiroPeriodo={meta.get('PrimeiroPeriodo')}")

        ok = meta.get("Periodic") == "Anual"
        (L.bate if ok else L.nao_bate)("periodicidade (INE)", "Anual", f"Periodic={meta.get('Periodic')}")

        ok = meta.get("UltimoPeriodo") == "2024 / 2025"
        (L.bate if ok else L.nao_bate)("ultimo_periodo (INE)", "2024 / 2025", f"UltimoPeriodo={meta.get('UltimoPeriodo')}")

        ok = meta.get("DataUltimaAtualizacao") == "2026-07-14"
        (L.bate if ok else L.nao_bate)("publicado_em (INE)", "2026-07-14", f"DataUltimaAtualizacao={meta.get('DataUltimaAtualizacao')}")

        niveis = niveis_geo(meta)
        ok = niveis.get("5") == 308
        (L.bate if ok else L.nao_bate)("concelho (INE)", "sim, 308 municipios", f"niveis={niveis}")

        t_lic = licenca_ine_texto()
        ok = "licença Creative Commons CC BY Atribuição 4.0" in t_lic
        (L.bate if ok else L.nao_bate)("licenca (INE)", "CC BY Atribuicao 4.0", f"confirmado={ok}")

        if data:
            vals = valores_por_geo(data, "2024 / 2025")
            ok = vals.get("PT") == "98,3" and vals.get("1") == "98,2" and vals.get("1C40705") == "107,0"
            (L.bate if ok else L.nao_bate)("excerto / valor_recente (INE)",
                                            "PT 98,3 ; Continente 98,2 ; Evora 107,0",
                                            f"PT={vals.get('PT')} ; Continente={vals.get('1')} ; Evora={vals.get('1C40705')}")
        else:
            L.na("excerto / valor_recente (INE)", "-", f"pindica falhou ({st_d})")

    # Eurostat educ_uoe_enra21
    url_eu = "https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/educ_uoe_enra21?format=JSON&lang=EN&geo=PT&geo=EU27_2020"
    st_eu, body_eu, err_eu = fetch(url_eu)
    if st_eu != 200:
        L.na("(parte Eurostat)", "-", f"pedido falhou ({st_eu}/{err_eu})")
        return L
    d = json.loads(body_eu)

    ok = d.get("label") == ("Pupils from age 3 to the starting age of compulsory education at primary level by "
                             "sex - % of the population of the corresponding age")
    (L.bate if ok else L.nao_bate)("excerto (label Eurostat)", "label conforme citado", f"label atual={d.get('label')!r}")

    ok = d.get("updated") == "2026-08-11T23:00:00+0200"
    (L.bate if ok else L.nao_bate)("publicado_em (Eurostat)", "2026-08-11T23:00:00+0200", f"updated={d.get('updated')}")

    anos = sorted(d["dimension"]["time"]["category"]["index"], key=lambda k: d["dimension"]["time"]["category"]["index"][k])
    ok = anos[0] == "2013" and anos[-1] == "2024"
    (L.bate if ok else L.nao_bate)("serie_desde / ultimo_periodo (Eurostat)", "2013 a 2024", f"anos={anos[0]}..{anos[-1]}")

    freq = list(d["dimension"]["freq"]["category"]["label"].values())[0]
    (L.bate if freq == "Annual" else L.nao_bate)("periodicidade (Eurostat)", "Anual (freq A)", f"freq={freq}")

    ids_ = d["id"]
    sizes = d["size"]
    strides = [1] * len(sizes)
    for i in range(len(sizes) - 2, -1, -1):
        strides[i] = strides[i + 1] * sizes[i + 1]
    idxmap = {ids_[i]: d["dimension"][ids_[i]]["category"]["index"] for i in range(len(ids_))}

    def pos(**kw):
        p = 0
        for i, dim in enumerate(ids_):
            chave = kw.get(dim, list(idxmap[dim].keys())[0])
            p += idxmap[dim][chave] * strides[i]
        return p

    vals = d["value"]
    status_flags = d.get("status", {})
    p_pt = pos(geo="PT", time="2024")
    p_eu = pos(geo="EU27_2020", time="2024")
    v_pt, f_pt = vals.get(str(p_pt)), status_flags.get(str(p_pt))
    v_eu, f_eu = vals.get(str(p_eu)), status_flags.get(str(p_eu))
    ok = v_pt == 94.5 and f_pt == "d" and v_eu == 94.9 and f_eu == "d"
    (L.bate if ok else L.nao_bate)("excerto / valor_recente (Eurostat)",
                                    "PT 2024: 94.5 (d) ; EU27_2020 2024: 94.9 (d)",
                                    f"PT={v_pt} ({f_pt}) ; EU27_2020={v_eu} ({f_eu})")

    st_flag, body_flag, _ = fetch("https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/"
                                   "educ_uoe_enra21?format=JSON&lang=EN&geo=PT&geo=EU27_2020")
    ok = b"definition differs (see metadata)" in body_flag
    (L.bate if ok else L.nao_bate)("excerto (significado da marca 'd')", "d = definition differs (see metadata)",
                                    f"confirmado={ok}")

    st_esms, body_esms, _ = fetch("https://ec.europa.eu/eurostat/cache/metadata/en/educ_uoe_enr_esms.htm")
    t_esms = texto_de_html(body_esms) if st_esms == 200 else ""
    frag_def = ("Pupils from age 0, 3, 4 and 5 to the starting age of compulsory education at primary level, "
                "as % of the population of the corresponding age. In some countries, the start of primary "
                "education is not compulsory and in some countries compulsory education starts at "
                "pre-primary level.")
    ok = frag_def in t_esms
    (L.bate if ok else L.nao_bate)("definicao (Eurostat)", frag_def[:80] + "...", f"confirmado no ESMS={ok}")

    st_semgeo, body_semgeo, _ = fetch("https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/"
                                       "educ_uoe_enra21?format=JSON&lang=EN&time=2024")
    if st_semgeo == 200:
        geo = json.loads(body_semgeo)["dimension"]["geo"]["category"]["index"]
        ok = len(geo) == 38 and "PT" in geo
        (L.bate if ok else L.nao_bate)("concelho (Eurostat, sem filtro geo)",
                                        "38 posicoes, unica portuguesa e PT",
                                        f"posicoes={len(geo)} ; PT presente={'PT' in geo}")
    else:
        L.na("concelho (Eurostat, sem filtro geo)", "-", f"pedido falhou ({st_semgeo})")

    st_lic_eu, body_lic_eu, _ = fetch("https://ec.europa.eu/eurostat/web/main/help/copyright-notice")
    t_lic_eu = texto_de_html(body_lic_eu) if st_lic_eu == 200 else ""
    frag_lic_eu = ("Reuse of statistical data, metadata, publications, and other dissemination tools published "
                   "on this website for commercial or non-commercial purposes is authorised provided the "
                   "source is acknowledged. The reuse policy of the European Commission is implemented by "
                   "the Decision of 12 December 2011")  # sem o ponto final: a conversao de HTML para texto
                   # insere um espaco antes do ponto nesta pagina ("2011 ."), artefacto irrelevante para o conteudo
    ok = frag_lic_eu in t_lic_eu
    (L.bate if ok else L.nao_bate)("licenca (Eurostat)", frag_lic_eu[:80] + "...", f"confirmado={ok}")
    L.na("licenca (DGEEC)", "[verify] (ver D2)", "a propria linha ja marca como [verify]; nao testado")

    return L


# ---------------------------------------------------------------------------
# H4 - Medicas/os por 1000 habitantes (INE 0012837)
# ---------------------------------------------------------------------------

def checar_H4(linha_json):
    L = Linha("H4")
    meta, _, st = ine_meta("0012837")
    if meta is None:
        L.na("(todas)", "-", f"pindicaMeta falhou ({st})")
        return L
    minfo, _, _ = ine_minfo("0012837")
    data, _, st_d = ine_data("0012837", "S7A2025", dim2="PT,1C40705")

    fonte_ok = "INE, Estatísticas do pessoal de saúde" in minfo
    (L.bate if fonte_ok else L.nao_bate)("publicador_primario", "Fonte: INE, Estatísticas do pessoal de saúde",
                                          f"confirmado={fonte_ok}")

    def_ok = ("Pessoa com pelo menos um diploma básico de medicina e que pratica ou praticou medicina" in minfo
              and "Fonte: Estatuto da Ordem dos Médicos, arts. 1.º e 8.º" in minfo)
    formula_ok = ("(Número total de médicas/os inscritos no final do ano/ População residente estimada para "
                  "o final do ano)*1000" in minfo)
    (L.bate if (def_ok and formula_ok) else L.nao_bate)(
        "definicao", "Definicao MEDICA/O verbatim + Formula verbatim", f"definicao={def_ok}; formula={formula_ok}")

    ok = meta.get("PrimeiroPeriodo") == "2021"
    (L.bate if ok else L.nao_bate)("serie_desde", "2021 (serie de cinco anos)", f"PrimeiroPeriodo={meta.get('PrimeiroPeriodo')}")

    ok = meta.get("Periodic") == "Anual"
    (L.bate if ok else L.nao_bate)("periodicidade", "Anual", f"Periodic={meta.get('Periodic')}")

    ok = meta.get("UltimoPeriodo") == "2025"
    (L.bate if ok else L.nao_bate)("ultimo_periodo", "2025", f"UltimoPeriodo={meta.get('UltimoPeriodo')}")

    ok = meta.get("DataUltimaAtualizacao") == "2026-07-28"
    (L.bate if ok else L.nao_bate)("publicado_em", "2026-07-28", f"DataUltimaAtualizacao={meta.get('DataUltimaAtualizacao')}")

    niveis = niveis_geo(meta)
    ok = niveis == {"1": 1, "2": 3, "3": 9, "4": 26, "5": 308}
    (L.bate if ok else L.nao_bate)("concelho", "sim, os 308 (1+3+9+26+308)", f"niveis={niveis}")

    t_lic = licenca_ine_texto()
    ok = "licença Creative Commons CC BY Atribuição 4.0" in t_lic
    (L.bate if ok else L.nao_bate)("licenca", "CC BY Atribuicao 4.0", f"confirmado={ok}")

    if data:
        vals = valores_por_geo(data, "2025")
        pt_ok = vals.get("PT") == "5,7"
        evora_fonte = vals.get("1C40705")
        evora_ok_excerto = evora_fonte == "8,3"       # o que a linha diz que a FONTE mostra (excerto)
        evora_ok_valor_recente = evora_fonte == "8,8"  # o que a linha afirma como valor mais recente

        (L.bate if pt_ok else L.nao_bate)("excerto (Portugal)", "PT 2025: 5,7", f"PT={vals.get('PT')}")
        (L.bate if evora_ok_excerto else L.nao_bate)("excerto (Evora)", "Evora 2025: 8,3 (citado no excerto)",
                                                       f"Evora={evora_fonte}")
        (L.bate if evora_ok_valor_recente else L.nao_bate)(
            "valor_recente (Evora)", "Evora 2025: 8,8 (afirmado em valor_recente)", f"Evora na fonte hoje={evora_fonte}")
    else:
        L.na("excerto / valor_recente", "-", f"pindica falhou ({st_d})")

    return L


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    inicio = time.time()
    with open(AMOSTRA_JSON, encoding="utf-8") as fh:
        amostra = {item["id"]: item for item in json.load(fh)}

    ordem = ["E3", "T3", "P1", "M2", "S3", "A1", "D3", "H4"]
    funcoes = {"E3": checar_E3, "T3": checar_T3, "P1": checar_P1, "M2": checar_M2,
               "S3": checar_S3, "A1": checar_A1, "D3": checar_D3, "H4": checar_H4}

    linhas_resultado = []
    for id_ in ordem:
        print(f"--- a verificar {id_} ---", file=sys.stderr)
        t0 = time.time()
        L = funcoes[id_](amostra[id_])
        linhas_resultado.append(L)
        print(f"    {id_} concluida em {time.time()-t0:.1f}s", file=sys.stderr)

    duracao = time.time() - inicio
    escrever_relatorio(linhas_resultado, duracao)
    print(f"\nRelatorio escrito em {RELATORIO_MD}", file=sys.stderr)
    print(f"Duracao total: {duracao/60:.1f} minutos", file=sys.stderr)
    print(f"Pedidos falhados na primeira tentativa: {len(falhas_log)}", file=sys.stderr)
    for f in falhas_log:
        print(f"  - {f}", file=sys.stderr)


def escrever_relatorio(linhas_resultado, duracao_seg):
    todas_nao_batem = []
    for L in linhas_resultado:
        for (celula, afirma, fonte, veredito) in L.resultados:
            if veredito == "NAO BATE":
                todas_nao_batem.append((L.id, celula, afirma, fonte))

    partes = []
    partes.append("# Medicao cega da amostra do inventario de fontes (01.09.2026)\n")
    partes.append("Modelo: Claude Sonnet 5\n")
    partes.append(
        "\nEste relatorio foi gerado pelo script `amostra-sonnet.py`, que contactou as fontes primarias "
        "de cada linha (INE, Eurostat, ERSAR) sem consultar o trabalho de quem escreveu a amostra, e "
        "comparou o que a fonte diz hoje com o que cada celula da linha afirma.\n")

    for L in linhas_resultado:
        partes.append(f"\n## Linha {L.id}\n")
        partes.append("| Celula | O que a linha afirma | O que a fonte diz hoje | Veredito |")
        partes.append("|---|---|---|---|")
        for (celula, afirma, fonte, veredito) in L.resultados:
            celula_e = celula.replace("|", "/")
            afirma_e = str(afirma).replace("|", "/").replace("\n", " ")[:200]
            fonte_e = str(fonte).replace("|", "/").replace("\n", " ")[:300]
            partes.append(f"| {celula_e} | {afirma_e} | {fonte_e} | {veredito} |")

    partes.append("\n## Celulas que nao batem\n")
    if todas_nao_batem:
        for (id_, celula, afirma, fonte) in todas_nao_batem:
            partes.append(f"- {id_} / {celula}: linha afirma {afirma!r}; fonte diz {fonte!r}")
    else:
        partes.append("- nenhuma (nesta execucao)")

    partes.append("\n## Pedidos que falharam na primeira tentativa\n")
    if falhas_log:
        for f in falhas_log:
            partes.append(f"- {f}")
    else:
        partes.append("- nenhum")

    partes.append(f"\n## Tempo gasto\n\n{duracao_seg/60:.1f} minutos ({duracao_seg:.0f} segundos).\n")

    with open(RELATORIO_MD, "w", encoding="utf-8") as fh:
        fh.write("\n".join(partes))


if __name__ == "__main__":
    main()
